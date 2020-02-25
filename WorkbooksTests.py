from openpyxl import load_workbook

def saveData(epochs,cost):
    Workbook = load_workbook('C:\CodingProjects\CreativeInquiry\MNIST-Neural\EpochsAndCost.xlsx')
    defaultSheet = Workbook['Sheet1']
    defaultSheet.cell(int(float(epochs/1000+1)),1,epochs)
    defaultSheet.cell(int(float(epochs/1000+1)),2,cost)
    Workbook.save('C:\CodingProjects\CreativeInquiry\MNIST-Neural\EpochsAndCost.xlsx')
